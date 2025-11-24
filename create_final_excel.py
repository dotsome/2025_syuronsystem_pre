#!/usr/bin/env python3
"""
最新のモデル比較結果をExcelファイルに出力（Mermaid画像付き）
model_comparison_detailed_20251120_171346.json 形式に対応
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

def create_comparison_excel(json_file: str, output_excel: str):
    """
    モデル比較結果をExcelファイルに出力
    """
    print("=" * 80)
    print("モデル比較Excelファイルを作成します")
    print("=" * 80)
    print()

    # JSONファイルを読み込み
    with open(json_file, 'r', encoding='utf-8') as f:
        results = json.load(f)

    print(f"📊 {len(results)}件のテスト結果を処理します\n")

    # Mermaid画像のディレクトリ
    mermaid_dir = Path(__file__).parent / "mermaid_outputs"

    # Excelワークブックを作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # 質問ごとにシートを作成
    questions = {}
    for result in results:
        q_id = result['question_id']
        if q_id not in questions:
            questions[q_id] = []
        questions[q_id].append(result)

    for q_id, q_results in sorted(questions.items()):
        ws = wb.create_sheet(title=q_id)

        # タイトル行
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"{q_id}: {q_results[0]['question']}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # ヘッダー行
        headers = [
            'Mermaidモデル', '回答モデル', 'Mermaid図(ラフ)', 'Mermaid図(整形後)', '回答', 'CSV出力',
            '登場人物判定(s)', '中心人物特定(s)', 'Mermaid生成(s)', 'CSV変換(s)', '回答生成(s)', '合計時間(s)'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # データ行
        current_row = 3
        for result in q_results:
            # モデル名
            ws.cell(row=current_row, column=1, value=result['mermaid_model'])
            ws.cell(row=current_row, column=2, value=result['answer_model'])

            # Mermaid画像（ラフ）
            mermaid_model = result['mermaid_model']
            answer_model = result['answer_model']
            rough_png = mermaid_dir / f"{q_id}_{mermaid_model}_{answer_model}_rough.png"
            final_png = mermaid_dir / f"{q_id}_{mermaid_model}_{answer_model}.png"

            max_height = 0

            # ラフMermaid画像
            if rough_png.exists():
                img = XLImage(str(rough_png))
                img.width = 300
                img.height = int(img.height * (300 / img.width))
                ws.add_image(img, f'C{current_row}')
                max_height = max(max_height, img.height)
            else:
                ws.cell(row=current_row, column=3, value="(画像なし)")

            # 整形後Mermaid画像
            if final_png.exists():
                img = XLImage(str(final_png))
                img.width = 300
                img.height = int(img.height * (300 / img.width))
                ws.add_image(img, f'D{current_row}')
                max_height = max(max_height, img.height)
            else:
                ws.cell(row=current_row, column=4, value="(画像なし)")

            # 行の高さを調整
            if max_height > 0:
                ws.row_dimensions[current_row].height = max_height * 0.75

            # 回答
            answer_text = result['outputs'].get('answer', '')
            ws.cell(row=current_row, column=5, value=answer_text)
            ws.cell(row=current_row, column=5).alignment = Alignment(wrap_text=True, vertical='top')

            # CSV出力
            csv_text = result['outputs'].get('csv', '')
            ws.cell(row=current_row, column=6, value=csv_text)
            ws.cell(row=current_row, column=6).alignment = Alignment(wrap_text=True, vertical='top')

            # 処理時間（各ステップを個別の列に）
            processes = result.get('processes', {})

            # 登場人物判定
            char_judgment = processes.get('character_judgment', {}).get('time', 0)
            ws.cell(row=current_row, column=7, value=round(char_judgment, 2))

            # 中心人物特定
            center_person = processes.get('center_person', {}).get('time', 0)
            ws.cell(row=current_row, column=8, value=round(center_person, 2))

            # Mermaid生成
            mermaid_gen = processes.get('mermaid_generation', {}).get('time', 0)
            ws.cell(row=current_row, column=9, value=round(mermaid_gen, 2))

            # CSV変換
            csv_conv = processes.get('csv_conversion', {}).get('time', 0)
            ws.cell(row=current_row, column=10, value=round(csv_conv, 2))

            # 回答生成
            answer_gen = processes.get('answer_generation', {}).get('time', 0)
            ws.cell(row=current_row, column=11, value=round(answer_gen, 2))

            # 合計時間
            total_time = result.get('total_time', 0)
            ws.cell(row=current_row, column=12, value=round(total_time, 2))

            current_row += 1

        # 列幅を調整
        ws.column_dimensions['A'].width = 15  # Mermaidモデル
        ws.column_dimensions['B'].width = 15  # 回答モデル
        ws.column_dimensions['C'].width = 45  # Mermaid図(ラフ)
        ws.column_dimensions['D'].width = 45  # Mermaid図(整形後)
        ws.column_dimensions['E'].width = 50  # 回答
        ws.column_dimensions['F'].width = 40  # CSV出力
        ws.column_dimensions['G'].width = 15  # 登場人物判定
        ws.column_dimensions['H'].width = 15  # 中心人物特定
        ws.column_dimensions['I'].width = 15  # Mermaid生成
        ws.column_dimensions['J'].width = 15  # CSV変換
        ws.column_dimensions['K'].width = 15  # 回答生成
        ws.column_dimensions['L'].width = 15  # 合計時間

        print(f"  ✅ シート '{q_id}' を作成しました ({len(q_results)}件)")

    # サマリーシートを作成
    ws_summary = wb.create_sheet(title="サマリー", index=0)

    # タイトル
    ws_summary.merge_cells('A1:I1')
    title_cell = ws_summary['A1']
    title_cell.value = "モデル比較結果サマリー"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 30

    # サマリーヘッダー
    summary_headers = [
        'Mermaidモデル', '回答モデル', 'テスト数',
        '登場人物判定(s)', '中心人物特定(s)', 'Mermaid生成(s)', 'CSV変換(s)', '回答生成(s)', '合計時間(s)'
    ]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # モデル組み合わせごとの統計を計算
    model_stats = {}
    for result in results:
        key = (result['mermaid_model'], result['answer_model'])
        if key not in model_stats:
            model_stats[key] = {
                'char_judgment': [],
                'center_person': [],
                'mermaid_gen': [],
                'csv_conv': [],
                'answer_gen': [],
                'total': []
            }

        processes = result.get('processes', {})
        model_stats[key]['char_judgment'].append(processes.get('character_judgment', {}).get('time', 0))
        model_stats[key]['center_person'].append(processes.get('center_person', {}).get('time', 0))
        model_stats[key]['mermaid_gen'].append(processes.get('mermaid_generation', {}).get('time', 0))
        model_stats[key]['csv_conv'].append(processes.get('csv_conversion', {}).get('time', 0))
        model_stats[key]['answer_gen'].append(processes.get('answer_generation', {}).get('time', 0))
        model_stats[key]['total'].append(result.get('total_time', 0))

    # サマリーデータを挿入
    current_row = 3
    for (mermaid_model, answer_model), stats in sorted(model_stats.items()):
        ws_summary.cell(row=current_row, column=1, value=mermaid_model)
        ws_summary.cell(row=current_row, column=2, value=answer_model)
        ws_summary.cell(row=current_row, column=3, value=len(stats['total']))

        # 各ステップの平均時間
        ws_summary.cell(row=current_row, column=4, value=round(sum(stats['char_judgment']) / len(stats['char_judgment']), 2))
        ws_summary.cell(row=current_row, column=5, value=round(sum(stats['center_person']) / len(stats['center_person']), 2))
        ws_summary.cell(row=current_row, column=6, value=round(sum(stats['mermaid_gen']) / len(stats['mermaid_gen']), 2))
        ws_summary.cell(row=current_row, column=7, value=round(sum(stats['csv_conv']) / len(stats['csv_conv']), 2))
        ws_summary.cell(row=current_row, column=8, value=round(sum(stats['answer_gen']) / len(stats['answer_gen']), 2))
        ws_summary.cell(row=current_row, column=9, value=round(sum(stats['total']) / len(stats['total']), 2))

        current_row += 1

    # 列幅を調整
    for col in range(1, 10):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    print(f"  ✅ サマリーシートを作成しました\n")

    # Excelファイルを保存
    wb.save(output_excel)
    print(f"✨ 完了! Excelファイルを保存しました: {output_excel}")
    print(f"   ファイルサイズ: {Path(output_excel).stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    import sys

    # コマンドライン引数から入力ファイルを取得
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    else:
        json_file = "model_comparison_detailed_20251120_171346.json"

    # 出力ファイル名を入力ファイル名から生成
    json_path = Path(json_file)
    timestamp = json_path.stem.replace('model_comparison_detailed_', '')
    output_excel = f"model_comparison_final_{timestamp}.xlsx"

    if not json_path.exists():
        print(f"❌ ファイルが見つかりません: {json_file}")
        exit(1)

    create_comparison_excel(json_file, output_excel)
