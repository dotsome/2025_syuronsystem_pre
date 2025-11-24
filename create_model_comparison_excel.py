#!/usr/bin/env python3
"""
モデル比較結果をExcelファイルに出力（Mermaid画像付き）
"""
import json
import subprocess
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

def convert_mermaid_to_png(mermaid_file: Path, output_file: Path) -> bool:
    """
    MermaidファイルをPNG画像に変換

    Args:
        mermaid_file: 入力Mermaidファイル
        output_file: 出力PNGファイル

    Returns:
        変換成功したかどうか
    """
    try:
        # Mermaidファイルを読み込み、コードブロックマーカーを削除
        with open(mermaid_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # ```mermaid と最初の ``` の間だけを抽出
        if '```mermaid' in content:
            # 最初の```mermaidを削除
            content = content.split('```mermaid', 1)[1]
            # 次の```までを抽出
            if '```' in content:
                content = content.split('```', 1)[0]

        # 先頭の改行を削除
        content = content.lstrip('\n')

        # <br/>を<br>に統一（一部のMermaidパーサーは<br/>を認識しない）
        content = content.replace('<br/>', '<br>')

        # 一時ファイルに保存
        temp_file = mermaid_file.with_suffix('.tmp.mmd')
        with open(temp_file, 'w', encoding='utf-8') as f:
            f.write(content)

        cmd = [
            'mmdc',
            '-i', str(temp_file),
            '-o', str(output_file),
            '-b', 'transparent',
            '-w', '800',
            '-H', '600'
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)

        # 一時ファイルを削除
        temp_file.unlink()

        if result.returncode == 0 and output_file.exists():
            print(f"  ✅ 画像変換成功: {output_file.name}")
            return True
        else:
            print(f"  ❌ 画像変換失敗: {mermaid_file.name}")
            if result.stderr:
                print(f"     エラー: {result.stderr[:200]}")
            return False
    except Exception as e:
        print(f"  ❌ 変換エラー: {mermaid_file.name} - {e}")
        return False

def create_comparison_excel(json_file: str, output_excel: str):
    """
    モデル比較結果をExcelファイルに出力

    Args:
        json_file: 入力JSONファイル
        output_excel: 出力Excelファイル
    """
    print("=" * 80)
    print("モデル比較Excelファイルを作成します")
    print("=" * 80)

    # JSONを読み込み
    with open(json_file, 'r', encoding='utf-8') as f:
        results = json.load(f)

    print(f"\n📊 {len(results)}件のテスト結果を処理します\n")

    # Mermaid画像ディレクトリを作成
    image_dir = Path("mermaid_images")
    image_dir.mkdir(exist_ok=True)

    # Mermaidファイルを画像に変換
    print("🖼️  Mermaid図を画像に変換中...")
    image_map = {}
    for result in results:
        mermaid_file_info = result.get('mermaid_file', {})
        relative_path = mermaid_file_info.get('relative_path', '')

        if not relative_path:
            continue

        mermaid_file = Path(relative_path)
        if not mermaid_file.exists():
            print(f"  ⚠️  ファイルが見つかりません: {mermaid_file}")
            continue

        # 画像ファイル名を生成
        png_file = image_dir / mermaid_file.with_suffix('.png').name

        # 変換
        success = convert_mermaid_to_png(mermaid_file, png_file)
        if success:
            key = f"{result['question_id']}_{result['mermaid_model']}_{result['answer_model']}"
            image_map[key] = png_file

    print(f"\n✅ {len(image_map)}個の画像を生成しました\n")

    # Excelワークブックを作成
    print("📝 Excelファイルを作成中...")
    wb = Workbook()

    # 質問ごとにシートを作成
    questions = {}
    for result in results:
        q_id = result['question_id']
        if q_id not in questions:
            questions[q_id] = {
                'question': result['question'],
                'results': []
            }
        questions[q_id]['results'].append(result)

    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 各質問ごとにシートを作成
    for q_id, q_data in sorted(questions.items()):
        ws = wb.create_sheet(title=q_id)

        # ヘッダー行
        ws['A1'] = f"質問: {q_data['question']}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')

        # カラムヘッダー
        headers = ['Mermaidモデル', '回答モデル', '合計時間(秒)', 'Mermaid図', '回答']
        ws.append([''] * len(headers))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # データ行
        current_row = 3
        for result in q_data['results']:
            mermaid_model = result['mermaid_model']
            answer_model = result['answer_model']
            total_time = result['total_time']
            answer = result['outputs']['answer']

            # データを挿入
            ws.cell(row=current_row, column=1, value=mermaid_model)
            ws.cell(row=current_row, column=2, value=answer_model)
            ws.cell(row=current_row, column=3, value=round(total_time, 2))

            # 回答テキスト
            answer_cell = ws.cell(row=current_row, column=5, value=answer)
            answer_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Mermaid画像を挿入
            key = f"{q_id}_{mermaid_model}_{answer_model}"
            if key in image_map:
                img_path = image_map[key]
                try:
                    img = XLImage(str(img_path))
                    # 画像サイズを調整（幅を400ピクセルに）
                    img.width = 400
                    img.height = int(img.height * (400 / img.width)) if img.width > 0 else 300

                    # 画像を配置（D列）
                    img.anchor = f'D{current_row}'
                    ws.add_image(img)

                    # 行の高さを画像に合わせて調整
                    ws.row_dimensions[current_row].height = img.height * 0.75
                except Exception as e:
                    print(f"  ⚠️  画像挿入エラー: {img_path} - {e}")
                    ws.cell(row=current_row, column=4, value="[画像エラー]")
            else:
                ws.cell(row=current_row, column=4, value="[画像なし]")

            current_row += 1

        # 列幅を調整
        ws.column_dimensions['A'].width = 20  # Mermaidモデル
        ws.column_dimensions['B'].width = 20  # 回答モデル
        ws.column_dimensions['C'].width = 15  # 合計時間
        ws.column_dimensions['D'].width = 60  # Mermaid図
        ws.column_dimensions['E'].width = 80  # 回答

        print(f"  ✅ シート '{q_id}' を作成しました ({len(q_data['results'])}件)")

    # サマリーシートを作成
    ws_summary = wb.create_sheet(title="サマリー", index=0)
    ws_summary['A1'] = "モデル比較実験サマリー"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:F1')

    # サマリーヘッダー
    summary_headers = ['Mermaidモデル', '回答モデル', '平均時間(秒)', '最小時間(秒)', '最大時間(秒)', 'テスト数']
    ws_summary.append([''] * len(summary_headers))
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # モデル組み合わせごとの統計を計算
    model_stats = {}
    for result in results:
        key = (result['mermaid_model'], result['answer_model'])
        if key not in model_stats:
            model_stats[key] = []
        model_stats[key].append(result['total_time'])

    # サマリーデータを挿入
    current_row = 3
    for (mermaid_model, answer_model), times in sorted(model_stats.items()):
        ws_summary.cell(row=current_row, column=1, value=mermaid_model)
        ws_summary.cell(row=current_row, column=2, value=answer_model)
        ws_summary.cell(row=current_row, column=3, value=round(sum(times) / len(times), 2))
        ws_summary.cell(row=current_row, column=4, value=round(min(times), 2))
        ws_summary.cell(row=current_row, column=5, value=round(max(times), 2))
        ws_summary.cell(row=current_row, column=6, value=len(times))
        current_row += 1

    # 列幅を調整
    for col in range(1, 7):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20

    print(f"  ✅ サマリーシートを作成しました\n")

    # Excelファイルを保存
    wb.save(output_excel)
    print(f"✨ 完了! Excelファイルを保存しました: {output_excel}")
    print(f"   ファイルサイズ: {Path(output_excel).stat().st_size / 1024:.1f} KB")

if __name__ == "__main__":
    import sys

    # コマンドライン引数から入力ファイルを取得
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    else:
        json_file = "model_comparison_detailed_20251120_171346.json"

    # 出力ファイル名を入力ファイル名から生成
    json_path = Path(json_file)
    timestamp = json_path.stem.split('_')[-2:]  # 日時部分を抽出
    output_excel = f"model_comparison_final_{'_'.join(timestamp)}.xlsx"

    if not json_path.exists():
        print(f"❌ ファイルが見つかりません: {json_file}")
        exit(1)

    create_comparison_excel(json_file, output_excel)
